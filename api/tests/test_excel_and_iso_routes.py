from __future__ import annotations

import datetime as dt
import io

import numpy as np
import openpyxl
import pandas as pd
import pytest
from fastapi.testclient import TestClient

from ai.groq_narrator import NullNarrator
from api import create_app
from api.dependencies import (
    get_auth_repository,
    get_fx_provider,
    get_narrator,
    get_repository,
    require_api_key,
)
from core.io.fx.static import StaticFxDataProvider
from db.memory import InMemoryAuthRepository, InMemorySimulationRepository

ORDER_DATE = dt.date(2026, 1, 15)


def _series(end_date: dt.date, n_days: int = 700, sigma: float = 0.12, seed: int = 42) -> pd.Series:
    rng = np.random.default_rng(seed)
    dates = pd.bdate_range(end=pd.Timestamp(end_date), periods=n_days)
    log_returns = rng.normal(0, sigma / np.sqrt(252), n_days)
    log_returns[0] = 0.0
    return pd.Series(10.0 * np.exp(np.cumsum(log_returns)), index=dates, dtype="float64")


@pytest.fixture
def client():
    app = create_app()
    provider = StaticFxDataProvider(_series(ORDER_DATE))
    app.dependency_overrides[get_fx_provider] = lambda: provider
    app.dependency_overrides[get_repository] = lambda: InMemorySimulationRepository()
    app.dependency_overrides[get_auth_repository] = lambda: InMemoryAuthRepository()
    app.dependency_overrides[get_narrator] = lambda: NullNarrator()
    app.dependency_overrides[require_api_key] = lambda: None
    with TestClient(app) as test_client:
        yield test_client


def _order_payload(**overrides):
    payload = {
        "amount_foreign": 100_000.0,
        "foreign_currency": "USD",
        "domestic_currency": "MAD",
        "order_date": ORDER_DATE.isoformat(),
        "delivery_date": (ORDER_DATE + dt.timedelta(days=90)).isoformat(),
        "target_margin_pct": 0.15,
        "domestic_rate": 0.03,
        "foreign_rate": 0.05,
        "n_simulations": 1_000,
        "seed": 7,
    }
    payload.update(overrides)
    return payload


def test_export_simulation_excel_endpoint(client):
    response = client.post("/api/v1/export/excel/simulation", json={"order": _order_payload()})
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        "attachment; filename=parity_simulation_USD_MAD.xlsx"
        in response.headers["content-disposition"]
    )

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Executive Summary" in wb.sheetnames
    assert "Instruments Comparison" in wb.sheetnames
    assert "Distribution Percentiles" in wb.sheetnames


def test_export_portfolio_excel_endpoint(client):
    orders = [
        _order_payload(foreign_currency="USD", amount_foreign=100_000.0),
        _order_payload(foreign_currency="EUR", amount_foreign=80_000.0),
    ]
    response = client.post(
        "/api/v1/export/excel/portfolio",
        json={"orders": orders, "seed": 42},
    )
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Portfolio Risk" in wb.sheetnames


def test_camt053_import_endpoint(client):
    sample_camt = """<?xml version="1.0" encoding="UTF-8"?>
    <Document xmlns="urn:iso:std:iso:20022:tech:xsd:camt.053.001.02">
      <BkToCstmrStmt>
        <Stmt>
          <Id>STMT-API-99</Id>
          <CreDtTm>2026-03-20T10:00:00Z</CreDtTm>
          <Acct>
            <Id><IBAN>FR7630006000011234567890189</IBAN></Id>
            <Ccy>EUR</Ccy>
          </Acct>
          <Ntry>
            <NtryRef>TX-REC-1</NtryRef>
            <Amt Ccy="USD">75000.00</Amt>
            <CdtDbtInd>CRDT</CdtDbtInd>
            <BookgDt><Dt>2026-03-18</Dt></BookgDt>
          </Ntry>
        </Stmt>
      </BkToCstmrStmt>
    </Document>
    """
    response = client.post(
        "/api/v1/connectors/iso20022/camt053/import",
        json={"xml_content": sample_camt},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["statement_id"] == "STMT-API-99"
    assert body["iban"] == "FR7630006000011234567890189"
    assert body["total_transactions"] == 1
    assert body["sales_records_count"] == 1
    assert body["purchase_records_count"] == 0
    assert body["currency_summary"]["USD"]["credits"] == 75000.0


def test_pain001_generate_endpoint(client):
    payload = {
        "initiating_party": "Treasury Department",
        "debtor_name": "Parity Enterprise",
        "debtor_iban": "FR7630006000011234567890189",
        "debtor_bic": "BNPAFRPPXXX",
        "transfers": [
            {
                "end_to_end_id": "E2E-100",
                "amount": 25000.0,
                "currency": "EUR",
                "creditor_name": "FX Desk Broker",
                "creditor_iban": "DE89370400440532013000",
                "creditor_bic": "DEUTDEDDFXX",
                "remittance_info": "Settlement of EUR/USD forward contract",
            }
        ],
    }
    response = client.post("/api/v1/connectors/iso20022/pain001/generate", json=payload)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/xml"
    assert "pain.001.001.03" in response.text
    assert "E2E-100" in response.text
    assert "25000.00" in response.text
