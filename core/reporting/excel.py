from __future__ import annotations

import datetime as dt
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models.models import SimulationResult
from core.models.portfolio import PortfolioResult

# Colors
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SECTION_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="64748B")
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")
REGULAR_FONT = Font(name="Calibri", size=11, color="1E293B")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

# Risk Level Colors
RISK_COLORS = {
    "Low": "DCFCE7",  # Light Green
    "Moderate": "FEF9C3",  # Light Yellow
    "High": "FEE2E2",  # Light Red
    "Critical": "FCA5A5",  # Stronger Red
}


def _autofit_columns(ws: Any) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if "\n" in val_str:
                val_str = max(val_str.split("\n"), key=len)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def generate_simulation_excel(result: SimulationResult) -> bytes:
    """Generate a multi-tab institutional Excel report for a single currency simulation."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Executive Summary & Risk KRI ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "PARITY — FX Risk & Hedging Analysis"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Report generated on {dt.datetime.now(dt.UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = SUBTITLE_FONT

    # Order Details Table
    order = result.order
    ws1["A4"] = "TRANSACTION OVERVIEW"
    ws1["A4"].font = SECTION_FONT
    ws1["A4"].fill = SECTION_FILL
    ws1.merge_cells("A4:B4")

    order_rows = [
        ("Currency Pair", f"{order.foreign_currency}/{order.domestic_currency}"),
        ("Exposure Amount", order.amount_foreign),
        ("Foreign Currency", order.foreign_currency),
        ("Domestic Currency", order.domestic_currency),
        ("Order Date", order.order_date.isoformat()),
        ("Delivery Date", order.delivery_date.isoformat()),
        ("Horizon (Days)", result.horizon_days),
        ("Spot Rate at Order Date", result.spot_rate_order_date),
        ("CIP Theoretical Forward Rate", result.expected_terminal_rate),
        ("Breakeven Exchange Rate (S*)", result.breakeven_rate),
        ("Budgeted Margin", result.budgeted_margin_pct),
        ("Target Margin Floor", order.min_acceptable_margin_pct),
    ]

    for idx, (label, val) in enumerate(order_rows, start=5):
        cell_a = ws1[f"A{idx}"]
        cell_b = ws1[f"B{idx}"]
        cell_a.value = label
        cell_a.font = BOLD_FONT
        cell_a.border = THIN_BORDER
        cell_b.value = val
        cell_b.font = REGULAR_FONT
        cell_b.border = THIN_BORDER

        if "Amount" in label:
            cell_b.number_format = "#,##0.00"
        elif "Rate" in label:
            cell_b.number_format = "0.0000"
        elif "Margin" in label or "Floor" in label:
            cell_b.number_format = "0.00%"

    # Key Risk Indicators (KRI) Table
    ws1["D4"] = "KEY RISK INDICATORS (KRI)"
    ws1["D4"].font = SECTION_FONT
    ws1["D4"].fill = SECTION_FILL
    ws1.merge_cells("D4:E4")

    kri_rows = [
        ("Vulnerability Score (0 - 100)", result.vulnerability_score),
        ("Risk Level", result.risk_level.value),
        ("Probability Below Floor", result.probability_margin_below_threshold),
        ("Expected Shortfall (CVaR 95%)", result.expected_shortfall_margin_pct),
        ("Annualized Volatility (σ)", result.annualized_volatility),
        ("Optimal Hedge Ratio", result.hedge.optimal_hedge_ratio),
        ("Hedged Margin at Optimal Ratio", result.hedge.optimal_hedge_cvar_margin_pct),
        ("CVaR Improvement vs Unhedged", result.hedge.cvar_improvement_pct),
    ]

    for idx, (label, val) in enumerate(kri_rows, start=5):
        cell_d = ws1[f"D{idx}"]
        cell_e = ws1[f"E{idx}"]
        cell_d.value = label
        cell_d.font = BOLD_FONT
        cell_d.border = THIN_BORDER
        cell_e.value = val
        cell_e.font = REGULAR_FONT
        cell_e.border = THIN_BORDER

        if label.startswith("Vulnerability"):
            cell_e.font = BOLD_FONT
            bg_color = RISK_COLORS.get(result.risk_level.value, "FFFFFF")
            cell_e.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        elif (
            "Probability" in label
            or "Shortfall" in label
            or "Volatility" in label
            or "Hedge Ratio" in label
            or "Margin" in label
        ):
            cell_e.number_format = "0.00%"
        elif "Drift" in label:
            cell_e.number_format = "0.0000"

    # Executive Recommendation Text
    ws1["A18"] = "EXECUTIVE RECOMMENDATION"
    ws1["A18"].font = SECTION_FONT
    ws1["A18"].fill = SECTION_FILL
    ws1.merge_cells("A18:E18")

    ws1["A19"] = result.recommendation
    ws1["A19"].font = REGULAR_FONT
    ws1["A19"].alignment = Alignment(wrap_text=True)
    ws1.merge_cells("A19:E20")

    _autofit_columns(ws1)

    # --- Sheet 2: Hedging Instruments Comparison ---
    ws2 = wb.create_sheet(title="Instruments Comparison")
    ws2.views.sheetView[0].showGridLines = True

    ws2["A1"] = "HEDGING STRATEGY BENCHMARK"
    ws2["A1"].font = TITLE_FONT

    headers = [
        "Instrument",
        "Strategy Description",
        "Upfront Premium",
        "Expected Margin",
        "Worst-Case Margin",
        "Best-Case Margin",
        "CVaR 95% Margin",
        "Prob. Below Floor",
    ]

    for c_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for r_idx, outcome in enumerate(result.instrument_comparison, start=4):
        row_vals = [
            outcome.instrument.value.upper(),
            outcome.description,
            outcome.upfront_premium_domestic,
            outcome.expected_margin_pct,
            outcome.worst_case_margin_pct,
            outcome.best_case_margin_pct,
            outcome.cvar_margin_pct,
            outcome.probability_below_threshold,
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if c_idx == 1:
                cell.font = BOLD_FONT
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 3:
                cell.number_format = "#,##0.00"
            elif c_idx >= 4:
                cell.number_format = "0.00%"

    _autofit_columns(ws2)

    # --- Sheet 3: Distribution & Percentiles ---
    ws3 = wb.create_sheet(title="Distribution Percentiles")
    ws3.views.sheetView[0].showGridLines = True

    ws3["A1"] = "SIMULATED DISTRIBUTION PERCENTILES"
    ws3["A1"].font = TITLE_FONT

    p_headers = ["Percentile", "Simulated FX Rate", "Resulting Margin %"]
    for c_idx, h in enumerate(p_headers, start=1):
        cell = ws3.cell(row=3, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    percentile_keys = ["P5", "P10", "P25", "P50", "P75", "P90", "P95"]
    for r_idx, p_key in enumerate(percentile_keys, start=4):
        rate_val = result.rate_percentiles.get(p_key, 0.0)
        margin_val = result.margin_pct_percentiles.get(p_key, 0.0)

        c1 = ws3.cell(row=r_idx, column=1, value=p_key)
        c1.font = BOLD_FONT
        c1.alignment = Alignment(horizontal="center")
        c1.border = THIN_BORDER

        c2 = ws3.cell(row=r_idx, column=2, value=rate_val)
        c2.font = REGULAR_FONT
        c2.number_format = "0.0000"
        c2.border = THIN_BORDER

        c3 = ws3.cell(row=r_idx, column=3, value=margin_val)
        c3.font = REGULAR_FONT
        c3.number_format = "0.00%"
        c3.border = THIN_BORDER

    _autofit_columns(ws3)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_portfolio_excel(result: PortfolioResult) -> bytes:
    """Generate an institutional Excel report for a multi-currency portfolio."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Portfolio Risk"
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "PARITY — Multi-Currency Portfolio Risk Report"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Report generated on {dt.datetime.now(dt.UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = SUBTITLE_FONT

    # Summary
    ws1["A4"] = "PORTFOLIO METRICS"
    ws1["A4"].font = SECTION_FONT
    ws1["A4"].fill = SECTION_FILL
    ws1.merge_cells("A4:B4")

    metrics = [
        ("Total Revenue (Domestic)", result.total_revenue_domestic),
        ("Total Budgeted Cost (Domestic)", result.total_budgeted_cost_domestic),
        ("Budgeted Margin", result.budgeted_margin_pct),
        ("Expected Margin", result.expected_margin_pct),
        ("Portfolio CVaR 95% Margin", result.cvar_margin_pct),
        ("Undiversified CVaR 95%", result.undiversified_cvar_margin_pct),
        ("Diversification Benefit", result.diversification_benefit_pct),
        ("Vulnerability Score", result.vulnerability_score),
        ("Risk Level", result.risk_level.value),
        ("Probability Below Floor", result.probability_below_threshold),
    ]

    for idx, (label, val) in enumerate(metrics, start=5):
        cA = ws1[f"A{idx}"]
        cB = ws1[f"B{idx}"]
        cA.value = label
        cA.font = BOLD_FONT
        cA.border = THIN_BORDER
        cB.value = val
        cB.font = REGULAR_FONT
        cB.border = THIN_BORDER

        if "Revenue" in label or "Cost" in label:
            cB.number_format = "#,##0.00"
        elif "Margin" in label or "CVaR" in label or "Benefit" in label or "Probability" in label:
            cB.number_format = "0.00%"
        elif label == "Vulnerability Score":
            cB.font = BOLD_FONT
            bg = RISK_COLORS.get(result.risk_level.value, "FFFFFF")
            cB.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    # Currency Risk Contributions Table
    ws1["D4"] = "CURRENCY RISK CONTRIBUTIONS (EULER ALLOCATION)"
    ws1["D4"].font = SECTION_FONT
    ws1["D4"].fill = SECTION_FILL
    ws1.merge_cells("D4:G4")

    headers = ["Currency", "Net Exposure", "Component CVaR %", "Risk Contribution %"]
    for c_idx, h in enumerate(headers, start=4):
        cell = ws1.cell(row=5, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for r_idx, contrib in enumerate(result.risk_contributions, start=6):
        vals = [
            contrib.currency,
            contrib.net_exposure,
            contrib.component_cvar_margin_pct,
            contrib.contribution_pct,
        ]
        for c_idx, v in enumerate(vals, start=4):
            cell = ws1.cell(row=r_idx, column=c_idx, value=v)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if c_idx == 4:
                cell.font = BOLD_FONT
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 5:
                cell.number_format = "#,##0.00"
            elif c_idx >= 6:
                cell.number_format = "0.00%"

    _autofit_columns(ws1)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
