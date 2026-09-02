from __future__ import annotations

import datetime as dt
import io

import numpy as np
import openpyxl
import pandas as pd
import pytest

from core.app.risk_engine import MarginRiskEngine
from core.io.fx.static import StaticFxDataProvider
from core.models.models import OrderInput
from core.models.portfolio import PortfolioPosition, simulate_portfolio
from core.reporting.excel import generate_portfolio_excel, generate_simulation_excel


def _make_sample_order() -> OrderInput:
    return OrderInput(
        amount_foreign=100_000.0,
        foreign_currency="USD",
        domestic_currency="EUR",
        order_date=dt.date(2026, 1, 1),
        delivery_date=dt.date(2026, 4, 1),
        target_margin_pct=0.15,
        min_acceptable_margin_pct=0.05,
        domestic_rate=0.02,
        foreign_rate=0.04,
        n_simulations=1000,
        lookback_days=180,
        seed=42,
    )


def test_generate_simulation_excel():
    order = _make_sample_order()
    dates = [order.order_date - dt.timedelta(days=i) for i in range(200)]
    series_dict = {pd.Timestamp(d): 0.90 + 0.001 * (i % 10) for i, d in enumerate(dates)}
    provider = StaticFxDataProvider(pd.Series(series_dict, dtype="float64"))
    engine = MarginRiskEngine(provider)
    result = engine.run(order)

    excel_bytes = generate_simulation_excel(result)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # Load back with openpyxl to verify integrity
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames
    assert "Executive Summary" in sheet_names
    assert "Instruments Comparison" in sheet_names
    assert "Distribution Percentiles" in sheet_names

    # Check Executive Summary
    ws1 = wb["Executive Summary"]
    assert "PARITY — FX Risk & Hedging Analysis" in str(ws1["A1"].value)
    assert ws1["B5"].value == "USD/EUR"
    assert ws1["B6"].value == 100_000.0

    # Check Instruments Comparison
    ws2 = wb["Instruments Comparison"]
    assert ws2["A3"].value == "Instrument"
    # Header + 5 instruments = rows 3 to 8
    instruments_found = [ws2.cell(row=r, column=1).value for r in range(4, 9)]
    assert "NONE" in instruments_found
    assert "FORWARD" in instruments_found
    assert "OPTION" in instruments_found
    assert "COLLAR" in instruments_found
    assert "PARTICIPATING_FORWARD" in instruments_found

    # Check Distribution Percentiles
    ws3 = wb["Distribution Percentiles"]
    p_labels = [ws3.cell(row=r, column=1).value for r in range(4, 11)]
    assert p_labels == ["P5", "P10", "P25", "P50", "P75", "P90", "P95"]


def test_generate_portfolio_excel():
    positions = [
        PortfolioPosition(
            label="pos_usd",
            foreign_currency="USD",
            amount_foreign=100_000.0,
            revenue_domestic=115_000.0,
            spot=0.92,
            sigma_annual=0.12,
            drift_annual=-0.02,
            horizon_years=0.5,
        ),
        PortfolioPosition(
            label="pos_gbp",
            foreign_currency="GBP",
            amount_foreign=50_000.0,
            revenue_domestic=65_000.0,
            spot=1.18,
            sigma_annual=0.10,
            drift_annual=0.01,
            horizon_years=0.5,
        ),
    ]
    corr = np.array([[1.0, 0.4], [0.4, 1.0]])
    result = simulate_portfolio(positions, corr, ["GBP", "USD"], n_sims=5000, seed=123)

    excel_bytes = generate_portfolio_excel(result)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Portfolio Risk" in wb.sheetnames
    ws = wb["Portfolio Risk"]
    assert "PARITY — Multi-Currency Portfolio Risk Report" in str(ws["A1"].value)
    assert ws["B5"].value == pytest.approx(180_000.0)  # total revenue (115k + 65k)
