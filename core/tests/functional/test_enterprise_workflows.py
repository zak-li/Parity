from __future__ import annotations

import datetime as dt
import io
import xml.etree.ElementTree as ET

import numpy as np
import openpyxl
import pandas as pd

from core.app.portfolio_engine import PortfolioRiskEngine
from core.app.risk_engine import MarginRiskEngine
from core.connectors.iso20022 import (
    Camt053Parser,
    CreditTransferInstruction,
    Pain001Generator,
)
from core.io.fx.static import StaticFxDataProvider
from core.models.enums import CopulaType, HedgeInstrument
from core.models.instruments import (
    garman_kohlhagen_call,
    garman_kohlhagen_put,
    participating_forward_strike,
)
from core.models.models import OrderInput
from core.models.smile import FxVolatilitySmile
from core.reporting.excel import generate_portfolio_excel, generate_simulation_excel


class MultiPairHistoricalProvider:
    def __init__(self, data_by_pair: dict[str, pd.Series]) -> None:
        self._data = data_by_pair

    def get_historical_series(
        self, base: str, quote: str, start: dt.date, end: dt.date
    ) -> pd.Series:
        key = f"{base.upper()}_{quote.upper()}"
        if key not in self._data:
            key = base.upper()
        series = self._data[key]
        mask = (series.index >= pd.Timestamp(start)) & (series.index <= pd.Timestamp(end))
        return series.loc[mask].copy()


def _generate_synthetic_fx(
    base_rate: float, sigma: float, end_date: dt.date, n_days: int = 250, seed: int = 42
) -> pd.Series:
    rng = np.random.default_rng(seed)
    dates = pd.bdate_range(end=pd.Timestamp(end_date), periods=n_days)
    n = len(dates)
    returns = rng.normal(0, sigma / np.sqrt(252), n)
    returns[0] = 0.0
    prices = base_rate * np.exp(np.cumsum(returns))
    return pd.Series(prices, index=dates, dtype="float64")


# ============================================================================
# WORKFLOW 1: End-to-End Corporate Treasury Lifecycle
# camt.053 ingestion -> Risk Engine -> 5-Instruments Benchmark -> Excel -> pain.001
# ============================================================================


def test_corporate_treasury_end_to_end_lifecycle():
    # 1. Ingest Bank Statement (camt.053)
    camt_xml = """<?xml version="1.0" encoding="UTF-8"?>
    <Document xmlns="urn:iso:std:iso:20022:tech:xsd:camt.053.001.08">
      <BkToCstmrStmt>
        <Stmt>
          <Id>STMT-2026-TREASURY-01</Id>
          <CreDtTm>2026-03-01T09:00:00Z</CreDtTm>
          <Acct>
            <Id><IBAN>MA64011000000000012345678901</IBAN></Id>
            <Ccy>MAD</Ccy>
          </Acct>
          <Bal>
            <Tp><CdOrPrtry><Cd>OPBD</Cd></CdOrPrtry></Tp>
            <Amt Ccy="MAD">5000000.00</Amt>
            <CdtDbtInd>CRDT</CdtDbtInd>
          </Bal>
          <Ntry>
            <NtryRef>IMPORT-PAYMENT-USD-01</NtryRef>
            <Amt Ccy="USD">150000.00</Amt>
            <CdtDbtInd>DBIT</CdtDbtInd>
            <BookgDt><Dt>2026-03-01</Dt></BookgDt>
            <ValDt><Dt>2026-06-01</Dt></ValDt>
            <NtryDtls>
              <TxDtls>
                <RltdPties><Cdtr><Nm>Industrial Equipment Global LLC</Nm></Cdtr></RltdPties>
                <RmtInf><Ustrd>PO #8849 Equipment Delivery June 2026</Ustrd></RmtInf>
              </TxDtls>
            </NtryDtls>
          </Ntry>
        </Stmt>
      </BkToCstmrStmt>
    </Document>
    """
    statements = Camt053Parser.parse_all(camt_xml)
    assert len(statements) == 1
    stmt = statements[0]
    assert stmt.iban == "MA64011000000000012345678901"
    purchases = stmt.to_purchase_records()
    assert len(purchases) == 1
    purchase = purchases[0]
    assert purchase.currency == "USD"
    assert purchase.amount == 150_000.0

    # 2. Configure OrderInput from the Extracted Bank Flow
    order_date = dt.date(2026, 3, 1)
    delivery_date = purchase.expected_delivery_date or dt.date(2026, 6, 1)
    order = OrderInput(
        amount_foreign=purchase.amount,
        foreign_currency=purchase.currency,
        domestic_currency="MAD",
        order_date=order_date,
        delivery_date=delivery_date,
        target_margin_pct=0.18,
        min_acceptable_margin_pct=0.05,
        domestic_rate=0.03,
        foreign_rate=0.05,
        n_simulations=10_000,
        lookback_days=200,
        seed=100,
    )

    # 3. Simulate with MarginRiskEngine
    series_usd = _generate_synthetic_fx(base_rate=10.0, sigma=0.14, end_date=order_date, seed=42)
    provider = StaticFxDataProvider(series_usd)
    engine = MarginRiskEngine(provider)
    result = engine.run(order)

    # Validate Engine Results
    assert result.vulnerability_score >= 0
    assert len(result.instrument_comparison) == 5
    instruments_map = {o.instrument: o for o in result.instrument_comparison}
    assert HedgeInstrument.PARTICIPATING_FORWARD in instruments_map
    par_fwd = instruments_map[HedgeInstrument.PARTICIPATING_FORWARD]
    assert par_fwd.upfront_premium_domestic == 0.0
    assert (
        par_fwd.worst_case_margin_pct > instruments_map[HedgeInstrument.NONE].worst_case_margin_pct
    )

    # 4. Generate Institutional Excel Report
    excel_bytes = generate_simulation_excel(result)
    assert len(excel_bytes) > 5000

    # Inspect Excel Sheet Contents
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Executive Summary" in wb.sheetnames
    assert "Instruments Comparison" in wb.sheetnames
    assert "Distribution Percentiles" in wb.sheetnames

    summary_ws = wb["Executive Summary"]
    assert summary_ws["B5"].value == "USD/MAD"
    assert summary_ws["B6"].value == 150_000.0
    assert summary_ws["E5"].value == result.vulnerability_score

    inst_ws = wb["Instruments Comparison"]
    assert inst_ws["A4"].value == "NONE"
    assert inst_ws["A8"].value == "PARTICIPATING_FORWARD"

    # 5. Generate ISO 20022 pain.001 Payment Instruction for Supplier Settlement
    transfers = [
        CreditTransferInstruction(
            end_to_end_id="E2E-ORDER-8849",
            amount=150_000.00,
            currency="USD",
            creditor_name="Industrial Equipment Global LLC",
            creditor_iban="US33CHAS0000001234567890",
            creditor_bic="CHASUS33XXX",
            execution_date=delivery_date,
            remittance_info="PO #8849 Equipment Delivery Settlement",
        )
    ]
    pain_xml = Pain001Generator.generate(
        initiating_party="Acme Treasury SARL",
        debtor_name="Acme Industrial Morocco",
        debtor_iban=stmt.iban,
        debtor_bic="BCECMAMCXXX",
        transfers=transfers,
        message_id="MSG-CORP-2026-001",
    )

    # Validate pain.001 XML structure
    root = ET.fromstring(pain_xml)
    for elem in root.iter():
        if "}" in elem.tag:
            elem.tag = elem.tag.split("}", 1)[1]

    assert root.find(".//MsgId").text == "MSG-CORP-2026-001"
    assert root.find(".//CtrlSum").text == "150000.00"
    assert root.find(".//NbOfTxs").text == "1"
    assert root.find(".//DbtrAcct/Id/IBAN").text == "MA64011000000000012345678901"
    assert root.find(".//Cdtr/Nm").text == "Industrial Equipment Global LLC"


# ============================================================================
# WORKFLOW 2: Multi-Currency Portfolio with Student-t Copula & Excel Export
# ============================================================================


def test_portfolio_student_t_copula_vs_gaussian_functional():
    end_date = dt.date(2026, 2, 1)
    provider = MultiPairHistoricalProvider(
        {
            "USD": _generate_synthetic_fx(10.0, 0.15, end_date, seed=1),
            "EUR": _generate_synthetic_fx(11.0, 0.12, end_date, seed=2),
            "GBP": _generate_synthetic_fx(13.0, 0.18, end_date, seed=3),
        }
    )

    orders = [
        OrderInput(
            amount_foreign=100_000.0,
            foreign_currency="USD",
            domestic_currency="MAD",
            order_date=end_date,
            delivery_date=end_date + dt.timedelta(days=90),
            target_margin_pct=0.15,
            domestic_rate=0.03,
            foreign_rate=0.05,
            n_simulations=30_000,
            seed=42,
        ),
        OrderInput(
            amount_foreign=80_000.0,
            foreign_currency="EUR",
            domestic_currency="MAD",
            order_date=end_date,
            delivery_date=end_date + dt.timedelta(days=90),
            target_margin_pct=0.15,
            domestic_rate=0.03,
            foreign_rate=0.03,
            n_simulations=30_000,
            seed=42,
        ),
        OrderInput(
            amount_foreign=50_000.0,
            foreign_currency="GBP",
            domestic_currency="MAD",
            order_date=end_date,
            delivery_date=end_date + dt.timedelta(days=90),
            target_margin_pct=0.15,
            domestic_rate=0.03,
            foreign_rate=0.04,
            n_simulations=30_000,
            seed=42,
        ),
    ]

    engine_gauss = PortfolioRiskEngine(provider, copula=CopulaType.GAUSSIAN)
    engine_student = PortfolioRiskEngine(provider, copula=CopulaType.STUDENT_T, copula_dof=3.0)

    res_gauss = engine_gauss.run(orders, seed=42)
    res_student = engine_student.run(orders, seed=42)

    # Under fat-tailed joint devaluations (Student-t with nu=3), tail risk must be strictly heavier
    assert res_student.cvar_margin_pct < res_gauss.cvar_margin_pct
    assert res_student.cvar_margin_domestic < res_gauss.cvar_margin_domestic
    assert len(res_student.risk_contributions) == 3

    # Generate Excel Report for the Portfolio
    excel_bytes = generate_portfolio_excel(res_student)
    assert len(excel_bytes) > 4000
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Portfolio Risk"]
    # Check that all 3 currencies are listed in the Euler risk contribution table
    currencies_in_excel = [ws.cell(row=r, column=4).value for r in range(6, 9)]
    assert set(currencies_in_excel) == {"EUR", "GBP", "USD"}


# ============================================================================
# WORKFLOW 3: Malz FX Smile Implied Volatility Surface & Pricing Invariants
# ============================================================================


def test_smile_arbitrage_and_invariants():
    atm = 0.12
    rr = 0.035  # Calls more expensive than Puts (positive skew)
    bf = 0.010  # Pronounced wings curvature
    smile = FxVolatilitySmile(atm_vol=atm, rr_25=rr, bf_25=bf)

    spot, rd, rf, t = 1.0, 0.02, 0.03, 0.5
    pricer = smile.build_pricer(spot, rd, rf, t)

    # 1. Monotonicity of Call and Put prices with strike
    strikes = [0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15]
    calls = [pricer("call", k) for k in strikes]
    puts = [pricer("put", k) for k in strikes]

    # Calls must strictly decrease with strike
    for i in range(len(calls) - 1):
        assert calls[i] > calls[i + 1]

    # Puts must strictly increase with strike
    for i in range(len(puts) - 1):
        assert puts[i] < puts[i + 1]

    # 2. Positive option values
    assert all(c > 0 for c in calls)
    assert all(p > 0 for p in puts)

    # 3. Volatility clipping for extreme deltas
    assert smile.vol_by_delta(0.0001) == smile.vol_by_delta(0.01)
    assert smile.vol_by_delta(0.9999) == smile.vol_by_delta(0.99)


# ============================================================================
# WORKFLOW 4: Participating Forward Monotonicity & Invariants
# ============================================================================


def test_participating_forward_monotonicity():
    spot = 1.05
    rd = 0.02
    rf = 0.04
    sigma = 0.12
    t = 0.5
    forward = spot * np.exp((rd - rf) * t)

    # Higher participation alpha -> higher protective strike K*
    alphas = [0.20, 0.40, 0.60, 0.80]
    strikes = [
        participating_forward_strike(spot, forward, rd, rf, sigma, t, alpha) for alpha in alphas
    ]

    for i in range(len(strikes) - 1):
        assert strikes[i] < strikes[i + 1]

    # For each strike, check that garman_kohlhagen call matches (1 - alpha) * put
    for alpha, k_star in zip(alphas, strikes, strict=False):
        c = garman_kohlhagen_call(spot, k_star, rd, rf, sigma, t)
        p = garman_kohlhagen_put(spot, k_star, rd, rf, sigma, t)
        assert abs(c - (1.0 - alpha) * p) < 1e-6
